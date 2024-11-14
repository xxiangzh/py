#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import warnings
from tkinter import *
from tkinter import messagebox

from openpyxl import Workbook
from openpyxl import load_workbook


def write(dir, rows):
    # 创建excel文件对象
    w_wb = Workbook()
    # 获取第一个sheet
    w_sheet = w_wb.active
    for row in rows:
        w_sheet.append(row)
    # 保存
    w_wb.save(dir + "\\新文件.xlsx")


def read(path, keywords):
    # 解决 warn("Workbook contains no default style, apply openpyxl's default")
    warnings.simplefilter('ignore')

    # 读取excel文件数据
    r_wb = load_workbook(path, data_only=True)
    # 获取第一个sheet
    r_sheet = r_wb.active
    contain_rows = []
    for row in r_sheet.rows:
        is_contain = False
        row_data = []
        for cell in row:
            row_data.append(cell.value)
            try:
                if keywords in cell.value:
                    is_contain = True
            except:
                pass
        if is_contain:
            contain_rows.append(row_data)
    r_wb.close()
    return contain_rows


def get_all_files(path):
    file_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_list.append(os.path.join(root, file))
    return file_list


def output():
    keywords = entry1.get()
    dir = entry2.get()
    print(keywords)
    print(dir)
    file_list = get_all_files(dir)
    rows = []
    for path in file_list:
        rows.extend(read(path, keywords))
    write(dir, rows)
    messagebox.showinfo("提示", "找到%d条" % (len(rows)))


# 程序主入口
if __name__ == "__main__":
    win = Tk()
    Label(win, text='关键字：').grid(row=0, column=0)
    entry1 = Entry(win)
    entry1.grid(row=0, column=1)

    Label(win, text='路径：').grid(row=1, column=0)
    entry2 = Entry(win)
    entry2.grid(row=1, column=1)

    Button(win, text='确认', command=output).grid(row=2, column=0, columnspan=2)
    win.mainloop()
