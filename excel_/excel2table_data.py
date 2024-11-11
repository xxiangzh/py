#!/usr/bin/env python
# -*- coding: utf-8 -*-

import mysql.connector
from openpyxl import load_workbook


def creat_table(table_name, title_list):
    # 建立数据库连接
    dbc = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        passwd="123456",
        database="xzh"
    )
    cursor = dbc.cursor()

    titles = []
    for t in title_list:
        titles.append("`" + t + "` VARCHAR(255)")
    ddl_sql = "CREATE TABLE `" + table_name + "` ( " + ', '.join(titles) + " ) "
    cursor.execute(ddl_sql)

    # 关闭游标和连接
    cursor.close()
    dbc.close()


def save_data(table_name, title_list, row_list):
    # 建立数据库连接
    dbc = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        passwd="123456",
        database="xzh"
    )
    cursor = dbc.cursor()

    titles = []
    s = []
    for t in title_list:
        titles.append("`" + t + "`")
        s.append("%s")
    insert_sql = "INSERT INTO " + table_name + " ( " + ', '.join(titles) + " ) VALUES ( " + ', '.join(s) + " )"
    # 逐条插入
    for val in row_list:
        cursor.execute(insert_sql, val)
        dbc.commit()

    # 关闭游标和连接
    cursor.close()
    dbc.close()


def read_title(path):
    # 读取excel文件数据
    wb = load_workbook(path, data_only=True)
    # 获取第一个sheet
    sheet = wb.active

    # 获取excel第一行标题
    row = next(sheet.rows)
    title_list = []
    for cell in row:
        title_list.append(cell.value)
    return title_list


def read_data(path):
    # 读取excel文件数据
    wb = load_workbook(path, data_only=True)
    # 获取第一个sheet
    sheet = wb.active

    # 将每行数据放到列表
    row_list = []
    for row in sheet.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        row_list.append(row_data)
    del row_list[0]
    return row_list


# 程序主入口
if __name__ == "__main__":
    path = "D:demo.xlsx"
    table_name = 't_user'
    title_list = read_title(path)
    creat_table(table_name, title_list)
    row_list = read_data(path)
    save_data(table_name, title_list, row_list)
    print("完成")
