#!/usr/bin/env python
# -*- coding: utf-8 -*-

import time

from openpyxl import Workbook
from openpyxl import load_workbook


def write():
    # 创建excel文件对象
    wb = Workbook()
    # 获取第一个sheet
    sheet = wb.active
    # 添加第1行标题
    sheet.append(['编号', '时间'])
    # 再添加5行数据
    for i in range(5):
        a = i + 1
        b = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        sheet.append([a, b])
    wb.save("D:demo.xlsx")


def read():
    # 读取excel文件数据
    wb = load_workbook('D:demo.xlsx', data_only=True)
    # 获取第一个sheet
    sheet = wb.active

    # 按行读取，返回A1, B1, C1顺序
    for row in sheet.rows:
        for cell in row:
            print(cell.value)

    # 按列读取，返回A1, A2, A3顺序
    for column in sheet.columns:
        for cell in column:
            print(cell.value)


# 程序主入口
if __name__ == "__main__":
    write()
    read()
