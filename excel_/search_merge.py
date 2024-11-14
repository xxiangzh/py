#!/usr/bin/env python
# -*- coding: utf-8 -*-
import warnings

from openpyxl import Workbook
from openpyxl import load_workbook


def write(rows):
    # 创建excel文件对象
    w_wb = Workbook()
    # 获取第一个sheet
    w_sheet = w_wb.active
    for row in rows:
        w_sheet.append(row)
    # 保存
    w_wb.save("D:新文件2.xlsx")


def read(path, keywords):
    # 解决 warn("Workbook contains no default style, apply openpyxl's default")
    warnings.simplefilter('ignore')

    # 读取excel文件数据
    r_wb = load_workbook(path, data_only=True)
    # 获取第一个sheet
    r_sheet = r_wb.active
    contain_rows = []
    for row in r_sheet.rows:
        is_contain = False
        row_data = []
        for cell in row:
            row_data.append(cell.value)
            try:
                if keywords in cell.value:
                    is_contain = True
            except:
                pass
        if is_contain:
            contain_rows.append(row_data)
    r_wb.close()
    return contain_rows


# 程序主入口
if __name__ == "__main__":
    rows1 = read('D:1.xlsx', '张三')
    rows2 = read('D:2.xlsx', '张三')
    rows = rows1 + rows2
    write(rows)
